# Leest excels in
# Excels hebben meerdere tabbladen
# Met een methode kan gekozen worden welk tabblad je wil doorlopen
# Standaard gaat de __iter__ over het eerste tabblad

from openpyxl import load_workbook


from extractor.reader import Reader
from model.entity import Header


class ExcelReader(Reader):
    def __init__(self, file_path):
        super().__init__(file_path)
        self.sheet_index = 0
        self.__name = None

    def set_name(self, name: str):
        super().set_name(name + '_' + str(self.sheet_index))

    def __iter__(self):
        wb = load_workbook(filename=self.file_path)
        sheet_name = wb.sheetnames[self.sheet_index]
        sheet = wb[sheet_name]

        iter_rows = iter(sheet.iter_rows())
        next(iter_rows)

        for row in iter_rows:
            row_values = [cell.value for cell in row]
            yield row_values

    def set_sheet_index(self, sheet_index):
        if 0 <= sheet_index < self.get_number_of_sheets():
            self.sheet_index = sheet_index
        else:
            raise IndexError("Invalid sheet index")

        # TODO fixed: INPUT VALIDEREN

    def __str__(self):
        return "Excel, eerste sheet met volgende headers = {self.get_header_names_per_sheet(0)}"

    def get_number_of_sheets(self) -> int:
        wb = load_workbook(filename=self.file_path)
        sheet_names = wb.sheetnames
        number_of_sheets = len(sheet_names)
        return number_of_sheets

    def get_header_info(self) -> [Header]:
        workbook = load_workbook(filename=self.file_path)
        sheet_name = workbook.sheetnames[self.sheet_index]
        sheet = workbook[sheet_name]
        excel_header_info = []

        # loop door de eerste rij: elke cell.value is de naam van de kolom
        # loop tegelijk door de tweede rij en het type(cell.value) is het datatype van de kolom
        # voeg de twee samen met 'header = Header(excel_header, datatype)'
        # En voeg ze toe aan de lijst excel_header_info en return die
        # TODO fixed: Header objecten terug geven
        # TODO fixed: gebruik de meta data van excel

        for cell_header, cell_data_type in zip(sheet[1], sheet[2]):
            excel_header = cell_header.value
            data_type = cell_data_type.data_type
            data_type_recoded = self.header_info_recoded(cell_data_type.value, data_type)
            header = Header(excel_header, data_type_recoded)
            excel_header_info.append(header)

        return excel_header_info

    def get_header_names(self):
        workbook = load_workbook(self.file_path)
        sheet_name = workbook.sheetnames[self.sheet_index]
        sheet = workbook[sheet_name]
        header_names = []
        for cell in sheet[1]:
            header_names.append(cell.value)
        return header_names

    def header_info_recoded(self, cell_value, cell_datatype):

        recode_dict = {
            's': 'TEXT',
            'd': 'DATETIME',  # 'DATETIME'
            'b': 'BOOLEAN',  # 'BOOLEAN'
            'e': 'TEXT'  # None?
        }
        if cell_datatype == 'n':
            try:
                int(cell_value)
                return 'INTEGER'
            except TypeError:
                try:
                    float(cell_value)
                    return 'FLOAT'
                except TypeError:
                    return 'TEXT'  # None?
        else:
            recoded_datatype = recode_dict.get(cell_datatype)
        return recoded_datatype

        # TODO fixed: header_info hercoderen (mag je er hier vanuit gaan dat de metadata correct is?)
        # wanneer ik de metadata van de excel file via openpyxl aanspreek (via de get_header_info methode)
        # krijg ik 's,n,b,d of e' terug. (string, numeric, boolean, date of error).
        # Deze moet ik nog even hercoderen naar waarden die voor Postgres acceptabel zijn
        # (string -> TEXT, numeric -> INTEGER of FLOAT, boolean -> TEXT?, date -> DATETIME, error -> raise Exception?)

    def __str__(self):
        return f"Excel, sheet_index {self.sheet_index} met volgende headers = {self.get_header_info()}"
